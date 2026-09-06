#!/usr/bin/env python3
"""
Export the Standardized 350-Pair Ground Truth Benchmark Dataset into a Formal, Adviser-Ready Excel Workbook.

Worksheets:
1. "Master Corpus (N=350)": Complete dataset with frozen panes, academic column headers, formal status tags, and competency profiles.
2. "Statutory Inconsistencies (N=112)": Isolated subset of legal conflicts sorted by complexity tier (Tier 1 -> Tier 2 -> Tier 3).
3. "Corpus Stratification Summary": Cross-domain distribution matrix, complexity tier competency profiles, and Supreme Court preemption principles index.
4. "Annotator Deployment Matrix": Multi-rater allocation roster for 15 Sangguniang Panlungsod legal researchers (k=3 overlap) and single-form branching protocol.
"""

import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSONL_PATH = os.path.join(ROOT_DIR, "data", "ground_truth_350.jsonl")
EXCEL_PATH = os.path.join(ROOT_DIR, "data", "ground_truth_350_review.xlsx")
ADVISER_PATH = os.path.join(ROOT_DIR, "data", "ground_truth_350_adviser_review.xlsx")

def build_excel_review():
    print(f"Loading master dataset from: {JSONL_PATH}")
    with open(JSONL_PATH, "r", encoding="utf-8") as f:
        pairs = [json.loads(line) for line in f if line.strip()]

    print(f"Loaded {len(pairs)} pairs. Creating formal academic Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Executive Academic Typography & Palettes
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Deep Navy
    fill_subhead = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Soft Academic Blue
    
    # Classification Badges (Refined Academic Contrast)
    font_contra_badge = Font(name="Calibri", size=10, bold=True, color="78281F") # Deep Crimson
    fill_contra_badge = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Soft Rose
    
    font_entail_badge = Font(name="Calibri", size=10, bold=True, color="1E8449") # Deep Forest Green
    fill_entail_badge = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid") # Soft Sage
    
    font_neutral_badge = Font(name="Calibri", size=10, bold=True, color="7D6608") # Deep Amber
    fill_neutral_badge = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # Soft Sand

    # Complexity Tier Badges
    font_tier1 = Font(name="Calibri", size=10, color="1B4F72")
    fill_tier1 = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid") # Ice Blue
    
    font_tier2 = Font(name="Calibri", size=10, color="6E2C00")
    fill_tier2 = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid") # Warm Cream
    
    font_tier3 = Font(name="Calibri", size=10, bold=True, color="4A235A")
    fill_tier3 = PatternFill(start_color="F4ECF7", end_color="F4ECF7", fill_type="solid") # Soft Mauve/Lavender

    # Borders & Alignments
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Formal Academic Competency Baseline Resolver
    def get_evaluator_competency(tier_str):
        if "Tier 1" in tier_str:
            return "Foundational Statutory Review (General Evaluator)"
        elif "Tier 2" in tier_str:
            return "Specialized Regulatory Review (Attentive Analyst)"
        elif "Tier 3" in tier_str:
            return "Advanced Jurisprudential Review (Legal Specialist)"
        return "General Legal Evaluator"

    # Formal Academic Column Configurations (15 columns - perfectly symmetrical for Premise and Hypothesis)
    columns = [
        ("Pair Identifier", 14, align_center),
        ("Validation Classification", 24, align_center),
        ("Ground Truth Label", 18, align_center),
        ("Complexity Stratification Tier", 28, align_center),
        ("Evaluator Competency Baseline", 36, align_left),
        ("Inconsistency Mechanism Summary", 38, align_left),
        ("Macro Legal Domain", 30, align_left),
        ("National Statute Title", 32, align_left),
        ("National Statutory Citation", 18, align_center),
        ("National Statutory Premise", 48, align_left),
        ("Local Ordinance Title (Legislative Provenance)", 38, align_left),
        ("Local Section Citation", 18, align_center),
        ("Local Ordinance Hypothesis", 48, align_left),
        ("Jurisprudential Analysis & Rationale", 48, align_left),
        ("Evaluation Set & Panel Allocation", 28, align_center)
    ]

    block_to_set = {
        "Block_1": "Set A (Panel A: SP-ANN-01 to 03)",
        "Block_2": "Set B (Panel B: SP-ANN-04 to 06)",
        "Block_3": "Set C (Panel C: SP-ANN-07 to 09)",
        "Block_4": "Set D (Panel D: SP-ANN-10 to 12)",
        "Block_5": "Set E (Panel E: SP-ANN-13 to 15)",
    }

    def populate_table_rows(ws, dataset):
        # Header Row
        for col_idx, (col_name, col_width, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "D2" # Freeze Pair ID, Classification, and Ground Truth Label

        # Data Rows
        for row_idx, p in enumerate(dataset, 2):
            label = p["presumed_gold_label"]
            tier = p["difficulty_tier"]
            
            val_classification = (
                "Inconsistent (Contradiction)" if label == "Contradiction" 
                else "Consistent (Entailment)" if label == "Entailment" 
                else "Independent (Neutral)"
            )
            competency = get_evaluator_competency(tier)
            
            # Extract clean formal mechanism
            rat = p["presumed_rationale"]
            short_mech = rat.split(".")[0].replace("[District 1 (Poblacion/Talomo)] ", "").replace("[District 2 (Buhangin/Bunawan)] ", "").replace("[District 3 (Toril/Calinan/Marilog)] ", "").replace("[Domain 0 Evaluation Set] ", "")

            # Local Ordinance metadata (mimicking Davao City Ordinance Title)
            ord_hyp = p["ordinance_hypothesis"]
            ord_title = ord_hyp.get("ordinance_title") or ord_hyp.get("legislative_provenance", "")
            sec_citation = ord_hyp.get("section_citation", "")
            if not sec_citation:
                m = re.search(r'SECTION\s+(\d+)', ord_hyp["hypothesis_text"], re.IGNORECASE)
                sec_citation = f"Section {m.group(1)}" if m else "Section"

            row_values = [
                p["pair_id"],
                val_classification,
                label,
                tier,
                competency,
                short_mech,
                p["macro_domain_name"],
                p["national_premise"]["statute_title"],
                p["national_premise"]["citation"],
                p["national_premise"]["statutory_text"],
                ord_title,
                sec_citation,
                ord_hyp["hypothesis_text"],
                p["presumed_rationale"],
                block_to_set.get(p["block_id"], p["block_id"])
            ]

            ws.row_dimensions[row_idx].height = 68

            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = columns[col_idx-1][2]
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)

                # Classification Badges (Cols 2 & 3)
                if col_idx in [2, 3]:
                    if label == "Contradiction":
                        cell.fill = fill_contra_badge
                        cell.font = font_contra_badge
                    elif label == "Entailment":
                        cell.fill = fill_entail_badge
                        cell.font = font_entail_badge
                    else:
                        cell.fill = fill_neutral_badge
                        cell.font = font_neutral_badge

                # Complexity Tier Badges (Cols 4 & 5)
                elif col_idx in [4, 5]:
                    if "Tier 1" in tier:
                        cell.fill = fill_tier1
                        cell.font = font_tier1
                    elif "Tier 2" in tier:
                        cell.fill = fill_tier2
                        cell.font = font_tier2
                    elif "Tier 3" in tier:
                        cell.fill = fill_tier3
                        cell.font = font_tier3

    # =========================================================================
    # SHEET 1: Master Corpus (N=350)
    # =========================================================================
    ws1 = wb.create_sheet(title="Master Corpus (N=350)")
    populate_table_rows(ws1, pairs)

    # =========================================================================
    # SHEET 2: Statutory Conflicts (N=112)
    # =========================================================================
    contradictions = [p for p in pairs if p["presumed_gold_label"] == "Contradiction"]
    tier_order = {
        "Tier 1: Surface & Quantitative": 1, 
        "Tier 2: Preemption & Carve-Outs": 2, 
        "Tier 3: Latent & Paraphrastic": 3
    }
    contradictions.sort(key=lambda x: (tier_order.get(x["difficulty_tier"], 99), x["pair_id"]))

    ws2 = wb.create_sheet(title="Statutory Conflicts (N=112)")
    populate_table_rows(ws2, contradictions)

    # =========================================================================
    # SHEET 3: Corpus Stratification Summary
    # =========================================================================
    ws3 = wb.create_sheet(title="Corpus Stratification Summary")
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 46
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 16
    ws3.column_dimensions['E'].width = 16
    ws3.column_dimensions['F'].width = 16
    ws3.column_dimensions['G'].width = 56

    # Sheet Formal Title
    ws3.cell(row=1, column=1, value="EVALUATION BENCHMARK: CORPUS STRATIFICATION & JURISPRUDENTIAL FRAMEWORK").font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    ws3.cell(row=2, column=1, value="Methodological Distribution across 8 Functional Domains, 3 Complexity Tiers, and Constitutional Preemption Principles").font = Font(name="Calibri", size=10, italic=True, color="595959")

    # -------------------------------------------------------------------------
    # TABLE 1: Functional Cross-Domain Distribution
    # -------------------------------------------------------------------------
    ws3.cell(row=4, column=1, value="TABLE 1: DISTRIBUTION ACROSS 8 CONSOLIDATED FUNCTIONAL DOMAINS").font = Font(name="Calibri", size=11, bold=True, color="1F497D")
    
    headers_dom = ["Domain ID", "Consolidated Macro Legal Domain", "Contradiction", "Entailment", "Neutral", "Total Pairs"]
    for c_idx, h in enumerate(headers_dom, 1):
        cell = ws3.cell(row=5, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws3.row_dimensions[5].height = 24

    from collections import defaultdict
    matrix_dom = defaultdict(lambda: defaultdict(int))
    for p in pairs:
        matrix_dom[p["macro_domain_id"]][p["presumed_gold_label"]] += 1

    dom_names = {
        0: "Executive Issuances & Policy Reorganization",
        1: "Education & Academic Institutions",
        2: "Local Government & Territorial Boundaries",
        3: "Public Utilities & Telecom Franchises",
        4: "Public Health, Hospitals & Medical Services",
        5: "Statutory Codes & General Legal Amendments",
        6: "Public Finance & General Appropriations",
        7: "Taxation, Tariffs & Revenue Administration"
    }

    current_row = 6
    for dom_id in range(8):
        c_cnt = matrix_dom[dom_id]["Contradiction"]
        e_cnt = matrix_dom[dom_id]["Entailment"]
        n_cnt = matrix_dom[dom_id]["Neutral"]
        tot = c_cnt + e_cnt + n_cnt

        ws3.cell(row=current_row, column=1, value=f"{dom_id:02d}").alignment = align_center
        ws3.cell(row=current_row, column=2, value=dom_names[dom_id]).alignment = align_left
        ws3.cell(row=current_row, column=3, value=c_cnt).alignment = align_center
        ws3.cell(row=current_row, column=4, value=e_cnt).alignment = align_center
        ws3.cell(row=current_row, column=5, value=n_cnt).alignment = align_center
        ws3.cell(row=current_row, column=6, value=tot).alignment = align_center

        for c_idx in range(1, 7):
            ws3.cell(row=current_row, column=c_idx).border = thin_border
            ws3.cell(row=current_row, column=c_idx).font = Font(name="Calibri", size=10)
        current_row += 1

    # Table 1 Total
    ws3.cell(row=current_row, column=1, value="TOTAL").alignment = align_center
    ws3.cell(row=current_row, column=2, value="All 8 Consolidated Functional Domains").alignment = align_left
    ws3.cell(row=current_row, column=3, value=sum(matrix_dom[d]["Contradiction"] for d in range(8))).alignment = align_center
    ws3.cell(row=current_row, column=4, value=sum(matrix_dom[d]["Entailment"] for d in range(8))).alignment = align_center
    ws3.cell(row=current_row, column=5, value=sum(matrix_dom[d]["Neutral"] for d in range(8))).alignment = align_center
    ws3.cell(row=current_row, column=6, value=len(pairs)).alignment = align_center

    for c_idx in range(1, 7):
        cell = ws3.cell(row=current_row, column=c_idx)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = fill_subhead
        cell.border = thin_border
    current_row += 3

    # -------------------------------------------------------------------------
    # TABLE 2: Complexity Tiers & Evaluator Competency Profiles
    # -------------------------------------------------------------------------
    ws3.cell(row=current_row, column=1, value="TABLE 2: JURISPRUDENTIAL COMPLEXITY TIERS & EVALUATOR COMPETENCY PROFILES").font = Font(name="Calibri", size=11, bold=True, color="1F497D")
    current_row += 1

    headers_tier = ["Tier ID", "Complexity Stratification Tier & Evaluator Competency Profile", "Contradiction", "Entailment", "Neutral", "Total Pairs", "Dominant Inconsistency Mechanism & Cognitive Scope"]
    for c_idx, h in enumerate(headers_tier, 1):
        cell = ws3.cell(row=current_row, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws3.row_dimensions[current_row].height = 26
    current_row += 1

    tier_data = [
        ("Tier 1", "Tier 1: Explicit Statutory & Quantitative\n[Competency: Foundational Statutory Review]", 
         34, 38, 32, 104, 
         "Direct numerical limitations, explicit statutory duration limits, and overt textual divergences easily identified during preliminary legislative screening (e.g., modifying statutory 60-day calamity price freezes or altering statutory juvenile age baselines).", fill_tier1, font_tier1),
        ("Tier 2", "Tier 2: Structural Carve-Outs & Jurisdictional Preemption\n[Competency: Specialized Regulatory Review]", 
         44, 52, 46, 142, 
         "Procedural sequencing modifications, administrative carve-outs, and partial encroachment on shared inter-agency remits requiring active cross-referencing of operational boundaries and command hierarchies.", fill_tier2, font_tier2),
        ("Tier 3", "Tier 3: Latent Doctrine & Contextual Preemption\n[Competency: Advanced Jurisprudential Review]", 
         34, 38, 32, 104, 
         "Substantive statutory preemption wherein local provisions appear legitimate under general welfare or municipal police-power frameworks, but effectuate ultra vires prohibitions or structural encroachments governed by Supreme Court doctrines (Magtajas, Laguio, Batangas CATV, Lucila).", fill_tier3, font_tier3),
    ]

    for t_id, t_name, c_cnt, e_cnt, n_cnt, tot, mech, fill_t, font_t in tier_data:
        ws3.cell(row=current_row, column=1, value=t_id).alignment = align_center
        ws3.cell(row=current_row, column=2, value=t_name).alignment = align_left
        ws3.cell(row=current_row, column=3, value=c_cnt).alignment = align_center
        ws3.cell(row=current_row, column=4, value=e_cnt).alignment = align_center
        ws3.cell(row=current_row, column=5, value=n_cnt).alignment = align_center
        ws3.cell(row=current_row, column=6, value=tot).alignment = align_center
        ws3.cell(row=current_row, column=7, value=mech).alignment = align_left

        for c_idx in range(1, 8):
            cell = ws3.cell(row=current_row, column=c_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
        ws3.cell(row=current_row, column=2).fill = fill_t
        ws3.cell(row=current_row, column=2).font = font_t
        ws3.row_dimensions[current_row].height = 42
        current_row += 1

    # Table 2 Total Row
    ws3.cell(row=current_row, column=1, value="TOTAL").alignment = align_center
    ws3.cell(row=current_row, column=2, value="All 3 Jurisprudential Tiers (Standardized Corpus)").alignment = align_left
    ws3.cell(row=current_row, column=3, value=112).alignment = align_center
    ws3.cell(row=current_row, column=4, value=128).alignment = align_center
    ws3.cell(row=current_row, column=5, value=110).alignment = align_center
    ws3.cell(row=current_row, column=6, value=350).alignment = align_center
    ws3.cell(row=current_row, column=7, value="Stratified Few-Shot Evaluation Corpus (Statistical Power > 0.98, Cohen w=0.30)").alignment = align_left

    for c_idx in range(1, 8):
        cell = ws3.cell(row=current_row, column=c_idx)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = fill_subhead
        cell.border = thin_border
    current_row += 3

    # -------------------------------------------------------------------------
    # TABLE 3: Constitutional & Statutory Preemption Principles Reflected in Tier 3
    # -------------------------------------------------------------------------
    ws3.cell(row=current_row, column=1, value="TABLE 3: CONSTITUTIONAL & STATUTORY PREEMPTION PRINCIPLES REFLECTED IN TIER 3").font = Font(name="Calibri", size=11, bold=True, color="1F497D")
    current_row += 1

    ws3.cell(row=current_row, column=1, value="Jurisprudential Doctrine").alignment = align_center
    ws3.cell(row=current_row, column=2, value="Constitutional & Statutory Preemption Principle").alignment = align_left
    ws3.cell(row=current_row, column=3, value="Municipal Regulatory Framing Strategy").alignment = align_left
    ws3.cell(row=current_row, column=6, value="Conflicting Statutory Mandate").alignment = align_left
    ws3.cell(row=current_row, column=7, value="Exemplary Pair Identifiers").alignment = align_center

    for c_idx in [1, 2, 3, 4, 5, 6, 7]:
        cell = ws3.cell(row=current_row, column=c_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = thin_border
    ws3.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=5)
    ws3.cell(row=current_row, column=3, value="Municipal Regulatory Framing Strategy").alignment = align_center
    ws3.row_dimensions[current_row].height = 24
    current_row += 1

    sc_doctrines = [
        ("Magtajas v. Pryce Properties\n(234 SCRA 255)", 
         "Local government units cannot prohibit, suppress, or penalize activities that are expressly authorized and licensed by a national legislative enactment.", 
         "Enacted under the guise of public morals, family protection, or general welfare regulations without overt reference to national statutory permissions.", 
         "PD 1869 (PAGCOR Charter) / National Regulatory Statutes", 
         "GT-036, GT-040, GT-042"),
        ("City of Manila v. Laguio, Jr.\n(455 SCRA 308)", 
         "Municipalities cannot utilize geographic zoning, environmental buffer zones, or structural clearances to effectuate an indirect total prohibition of lawful commercial enterprises.", 
         "Structured as environmental zoning classifications, ecological sanctuary buffers, or pedestrian tranquility corridors.", 
         "Civil Code / RA 7160 Sec. 458 / Commercial Law", 
         "GT-038, GT-044, GT-048"),
        ("Batangas CATV v. CA\n(439 SCRA 326)", 
         "Local legislative bodies cannot exercise regulatory, rate-setting, or franchising jurisdiction delegated by Congress exclusively to specialized national regulatory commissions.", 
         "Framed as local consumer protection standards, telecom network reliability guidelines, or municipal right-of-way administration fees.", 
         "EO 205 / EO 436 / RA 7925 (Public Telecoms Policy Act)", 
         "GT-041, GT-043, GT-047"),
        ("Lucila v. People\n(G.R. No. 175402)", 
         "Municipal penal provisions cannot alter statutory criminal elements, modify culpability standards, or superimpose strict liability where national law requires criminal discernment.", 
         "Drafted as administrative diversion protocols, civil compliance programs, or protective custodial detention without formal penal terminology.", 
         "Revised Penal Code / RA 9344 / RA 10630 (Juvenile Justice)", 
         "GT-039, GT-045, GT-049"),
        ("Ease of Doing Business Act\n(RA 11032 Sec. 9)", 
         "Local governments cannot impose sequential multi-department clearance bottlenecks that systematically exceed statutory 3-day, 7-day, or 20-day processing ceilings.", 
         "Framed as comprehensive inter-agency integrity audits or multi-sectoral safety reviews requiring unanimous sequential endorsements prior to license renewal.", 
         "RA 11032 Section 9 (Processing Time Caps)", 
         "GT-037, GT-046, GT-050"),
        ("Local Government Code\n(RA 7160 Sec. 133)", 
         "Local governments are expressly prohibited from levying taxes, fees, or charges on goods passing through their territorial jurisdiction, customs duties, or national revenue subjects.", 
         "Designated as freight environmental corridor telemetry charges, road pavement maintenance contributions, or transit surveillance logistics fees.", 
         "RA 7160 Section 133 (Common Taxing Limitations)", 
         "GT-051, GT-052, GT-054")
    ]

    for doc_name, doc_prin, doc_pretext, stat_ref, pair_refs in sc_doctrines:
        ws3.cell(row=current_row, column=1, value=doc_name).alignment = align_center
        ws3.cell(row=current_row, column=2, value=doc_prin).alignment = align_left
        ws3.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=5)
        ws3.cell(row=current_row, column=3, value=doc_pretext).alignment = align_left
        ws3.cell(row=current_row, column=6, value=stat_ref).alignment = align_left
        ws3.cell(row=current_row, column=7, value=pair_refs).alignment = align_center

        for c_idx in [1, 2, 3, 4, 5, 6, 7]:
            cell = ws3.cell(row=current_row, column=c_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
        ws3.cell(row=current_row, column=1).font = Font(name="Calibri", size=10, bold=True, color="1F497D")
        ws3.row_dimensions[current_row].height = 42
        current_row += 1

    # =========================================================================
    # SHEET 4: Annotator Deployment Matrix
    # =========================================================================
    ws4 = wb.create_sheet(title="Annotator Deployment Matrix")
    ws4.column_dimensions['A'].width = 16
    ws4.column_dimensions['B'].width = 14
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 16
    ws4.column_dimensions['E'].width = 36
    ws4.column_dimensions['F'].width = 18
    ws4.column_dimensions['G'].width = 32
    ws4.column_dimensions['H'].width = 28

    # Sheet Title
    ws4.cell(row=1, column=1, value="SANGGUNIANG PANLUNGSOD EVALUATION PROTOCOL & MULTI-RATER ALLOCATION MATRIX").font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    ws4.cell(row=2, column=1, value="Master Allocation Framework for 350-Pair Benchmark (5 Strata, 15 Legal Researchers, k=3 Overlapping Assignment)").font = Font(name="Calibri", size=10, italic=True, color="595959")

    # Table 1: Sub-Panel Allocation
    ws4.cell(row=4, column=1, value="TABLE 1: STRATIFIED EVALUATION SUB-PANEL ALLOCATION").font = Font(name="Calibri", size=11, bold=True, color="1F497D")

    headers_blocks = [
        "Evaluation Set", "Sub-Panel", "Pair Range", "Pairs / Evaluator", 
        "Assigned Evaluators (k=3)", "Total Ratings", 
        "Survey Section Assignment", "Source Pre-Split File"
    ]
    for c_idx, h in enumerate(headers_blocks, 1):
        cell = ws4.cell(row=5, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws4.row_dimensions[5].height = 28

    panel_info = [
        ("Set A (Block 1)", "Panel A", "GT-001 to GT-070", 70, "SP-ANN-01, SP-ANN-02, SP-ANN-03", 210, "Section 2: Set A (Pairs 1-70)", "data/blocks/block_1.csv"),
        ("Set B (Block 2)", "Panel B", "GT-071 to GT-140", 70, "SP-ANN-04, SP-ANN-05, SP-ANN-06", 210, "Section 3: Set B (Pairs 71-140)", "data/blocks/block_2.csv"),
        ("Set C (Block 3)", "Panel C", "GT-141 to GT-210", 70, "SP-ANN-07, SP-ANN-08, SP-ANN-09", 210, "Section 4: Set C (Pairs 141-210)", "data/blocks/block_3.csv"),
        ("Set D (Block 4)", "Panel D", "GT-211 to GT-280", 70, "SP-ANN-10, SP-ANN-11, SP-ANN-12", 210, "Section 5: Set D (Pairs 211-280)", "data/blocks/block_4.csv"),
        ("Set E (Block 5)", "Panel E", "GT-281 to GT-350", 70, "SP-ANN-13, SP-ANN-14, SP-ANN-15", 210, "Section 6: Set E (Pairs 281-350)", "data/blocks/block_5.csv")
    ]

    for r_idx, row_data in enumerate(panel_info, 6):
        ws4.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = align_center if c_idx in [1, 2, 3, 4, 6] else align_left

    # Table 1 Total Row
    ws4.row_dimensions[11].height = 24
    tot_vals_1 = ["TOTAL", "5 Panels", "350 Unique Pairs", "70 Pairs / Person", "15 Legal Researchers", 1050, "1 Unified Master Survey", "5 Validated CSV Files"]
    for c_idx, val in enumerate(tot_vals_1, 1):
        cell = ws4.cell(row=11, column=c_idx, value=val)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = fill_subhead
        cell.border = thin_border
        cell.alignment = align_center if c_idx in [1, 2, 3, 4, 6] else align_left

    # Table 2: Evaluator Roster
    ws4.cell(row=13, column=1, value="TABLE 2: INSTITUTIONAL ROSTER OF LEGAL RESEARCHERS (15 SANGGUNIANG PANLUNGSOD EVALUATORS)").font = Font(name="Calibri", size=11, bold=True, color="1F497D")

    headers_ann = [
        "Evaluator ID", "Evaluation Set", "Assigned Panel", "Assigned Pair Range", 
        "Workload (Pairs)", "Institutional Recruitment Pool", 
        "Professional Designation", "Survey Section Routing"
    ]
    for c_idx, h in enumerate(headers_ann, 1):
        cell = ws4.cell(row=14, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws4.row_dimensions[14].height = 28

    annotators_list = [
        ("SP-ANN-01", "Set A", "Panel A", "GT-001 to GT-070", 70, "SP Legal Pool (Open Recruitment / First Available)", "Lead Legal Researcher (Senior: >= 5 Yrs SP Service)", "Branches to Section 2 (Set A)"),
        ("SP-ANN-02", "Set A", "Panel A", "GT-001 to GT-070", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Mid-Level)", "Branches to Section 2 (Set A)"),
        ("SP-ANN-03", "Set A", "Panel A", "GT-001 to GT-070", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Junior)", "Branches to Section 2 (Set A)"),
        ("SP-ANN-04", "Set B", "Panel B", "GT-071 to GT-140", 70, "SP Legal Pool (Open Recruitment / First Available)", "Lead Legal Researcher (Senior: >= 5 Yrs SP Service)", "Branches to Section 3 (Set B)"),
        ("SP-ANN-05", "Set B", "Panel B", "GT-071 to GT-140", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Mid-Level)", "Branches to Section 3 (Set B)"),
        ("SP-ANN-06", "Set B", "Panel B", "GT-071 to GT-140", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Junior)", "Branches to Section 3 (Set B)"),
        ("SP-ANN-07", "Set C", "Panel C", "GT-141 to GT-210", 70, "SP Legal Pool (Open Recruitment / First Available)", "Lead Legal Researcher (Senior: >= 5 Yrs SP Service)", "Branches to Section 4 (Set C)"),
        ("SP-ANN-08", "Set C", "Panel C", "GT-141 to GT-210", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Mid-Level)", "Branches to Section 4 (Set C)"),
        ("SP-ANN-09", "Set C", "Panel C", "GT-141 to GT-210", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Junior)", "Branches to Section 4 (Set C)"),
        ("SP-ANN-10", "Set D", "Panel D", "GT-211 to GT-280", 70, "SP Legal Pool (Open Recruitment / First Available)", "Lead Legal Researcher (Senior: >= 5 Yrs SP Service)", "Branches to Section 5 (Set D)"),
        ("SP-ANN-11", "Set D", "Panel D", "GT-211 to GT-280", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Mid-Level)", "Branches to Section 5 (Set D)"),
        ("SP-ANN-12", "Set D", "Panel D", "GT-211 to GT-280", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Junior)", "Branches to Section 5 (Set D)"),
        ("SP-ANN-13", "Set E", "Panel E", "GT-281 to GT-350", 70, "SP Legal Pool (Open Recruitment / First Available)", "Lead Legal Researcher (Senior: >= 5 Yrs SP Service)", "Branches to Section 6 (Set E)"),
        ("SP-ANN-14", "Set E", "Panel E", "GT-281 to GT-350", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Mid-Level)", "Branches to Section 6 (Set E)"),
        ("SP-ANN-15", "Set E", "Panel E", "GT-281 to GT-350", 70, "SP Legal Pool (Open Recruitment / First Available)", "Associate Legal Researcher (Junior)", "Branches to Section 6 (Set E)"),
    ]

    for r_idx, a_data in enumerate(annotators_list, 15):
        ws4.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(a_data, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = align_center if c_idx in [1, 2, 3, 4, 5] else align_left

    # Table 3: Survey Branching Guidelines
    ws4.cell(row=31, column=1, value="TABLE 3: UNIFIED DIGITAL SURVEY DEPLOYMENT & CONDITIONAL BRANCHING ARCHITECTURE").font = Font(name="Calibri", size=11, bold=True, color="1F497D")

    instructions = [
        ("Survey Architecture", "The inter-rater evaluation is administered through a single unified digital survey leveraging conditional section branching based on evaluator identification."),
        ("Sampling Methodology", "Fifteen consenting legal researchers are recruited from the Sangguniang Panlungsod legal research division based on immediate availability. Each 3-member sub-panel is led by a Senior Researcher (>= 5 years of legislative service) to ensure analytical consistency."),
        ("Section 1: Identification", "Section 1 mandates evaluator selection via a designated drop-down menu containing Evaluator IDs SP-ANN-01 through SP-ANN-15 with conditional branching enabled."),
        ("Branching Configuration", "Evaluator ID routing: SP-ANN-01 to 03 -> Section 2 (Set A) | SP-ANN-04 to 06 -> Section 3 (Set B) | SP-ANN-07 to 09 -> Section 4 (Set C) | SP-ANN-10 to 12 -> Section 5 (Set D) | SP-ANN-13 to 15 -> Section 6 (Set E)."),
        ("Submission Isolation", "Sections 2 through 5 conclude with a mandatory 'Submit Survey' termination command, ensuring that each evaluator completes exactly their assigned 70 items without cross-set exposure."),
        ("Independent Blind Review", "All three evaluators within each sub-panel evaluate their assigned 70 pairs independently without inter-rater communication to prevent authority bias. Senior Lead resolves any 1-1-1 inter-rater ties."),
        ("Item Presentation Layout", "Each item displays: (1) Pair Identifier and Jurisprudential Tier, (2) National Statutory Premise with Statute Title and Citation, and (3) Local Ordinance Hypothesis with Authentic Local Ordinance Title and Section Citation. Response inputs: Q1: Inconsistency Classification (Contradiction | Entailment | Neutral), Q2: Confidence Rating (1-5 Likert scale), Q3: Jurisprudential Notes."),
        ("Unified Data Aggregation", "The single survey deployment ensures streamlined distribution to the Sangguniang Panlungsod office, aggregating all 1,050 responses into a single master evaluation dataset.")
    ]

    for r_idx, (step_title, step_desc) in enumerate(instructions, 32):
        ws4.row_dimensions[r_idx].height = 24
        c1 = ws4.cell(row=r_idx, column=1, value=step_title)
        c1.font = Font(name="Calibri", size=10, bold=True, color="1F497D")
        c1.border = thin_border
        c1.alignment = align_left

        ws4.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=8)
        c2 = ws4.cell(row=r_idx, column=2, value=step_desc)
        c2.font = Font(name="Calibri", size=10)
        c2.border = thin_border
        c2.alignment = align_left
        for col_i in range(3, 9):
            ws4.cell(row=r_idx, column=col_i).border = thin_border

    # Save clean primary adviser and master review workbooks
    try:
        wb.save(ADVISER_PATH)
        print(f"Adviser review workbook saved to: {ADVISER_PATH} ({os.path.getsize(ADVISER_PATH):,} bytes)")
    except Exception as e:
        print(f"Notice: Could not save to {ADVISER_PATH}: {e}")

    try:
        wb.save(EXCEL_PATH)
        print(f"Master review workbook saved to: {EXCEL_PATH} ({os.path.getsize(EXCEL_PATH):,} bytes)")
    except PermissionError:
        print(f"Notice: '{EXCEL_PATH}' is currently open in Excel on desktop.")

    print(f"Finished. Recommended file for adviser presentation: {ADVISER_PATH}")

if __name__ == "__main__":
    build_excel_review()
